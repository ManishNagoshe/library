from http import cookies
from unicodedata import decimal
from app import login
from fastapi import HTTPException, status,APIRouter
from jose import JWTError, jwt
from datetime import date

from pydantic import BaseModel
from app.config import settings
from typing import List, Optional
from fastapi import FastAPI,Cookie
import psycopg2
from psycopg2.extras import RealDictCursor
from starlette.responses import Response
from fpdf import FPDF
from openpyxl import Workbook
from starlette.responses import FileResponse

router=APIRouter(
    prefix="/report"
)

class Useridreport(BaseModel):
    id:int
    startindex:int
    endindex:int
@router.post("/user_report")
def userreport(response:Response,user:Useridreport,Manish:Optional[str]=Cookie(None)):
    role=login.getrole(Manish)
    if role!=1:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"invalid user")
    if(user.endindex-user.startindex>10):
        return({"msg":"Maximum 10 books can be fetched"})
    if(user.startindex<0):
        return({"msg":"Start index minimum value should be zero"})
    
    if(user.endindex<user.startindex):
        return({"msg":"end index should be greater than start index"})

    try:

        mydb = psycopg2.connect(
        host=settings.HOST_NAME,
        user=settings.USER_NAME,
        password=settings.USER_PASSWORD,
        database=settings.DATABASE_NAME,
        cursor_factory=RealDictCursor
        )
        mycursor = mydb.cursor()
        max_num=user.endindex-user.startindex
        val = (user.id,max_num,user.startindex)
        sql="SELECT title,authors,issuedate,returndate FROM usage join bookmaster on usage.accno=bookmaster.accno where userid=%s order by issuedate limit %s offset %s"
        mycursor.execute(sql, val,)
        myresult = mycursor.fetchall()
        mycursor.close()
        mydb.close()
        response.set_cookie(key="Manish",value=login.setcookie(login.verifyuser(Manish)), httponly=True,secure=settings.SECURITYHHTPS, samesite=settings.SAMESITE)
        if myresult:
            return(myresult)
        else:
            return
    except:
        return({"msg":"Connection error"})

class Personalreport(BaseModel):
    startindex:int
    endindex:int

@router.post("/personal_report")
def personalreport(response:Response,user:Personalreport,Manish:Optional[str]=Cookie(None)):
    role=login.getrole(Manish)
    userid=login.getuserid(Manish)
    if role!=2:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"invalid user")
    if(user.endindex-user.startindex>10):
        return({"msg":"Maximum 10 books can be fetched"})
    if(user.startindex<0):
        return({"msg":"Start index minimum value should be zero"})
    
    if(user.endindex<user.startindex):
        return({"msg":"end index should be greater than start index"})

    try:

        mydb = psycopg2.connect(
        host=settings.HOST_NAME,
        user=settings.USER_NAME,
        password=settings.USER_PASSWORD,
        database=settings.DATABASE_NAME,
        cursor_factory=RealDictCursor
        )
        mycursor = mydb.cursor()
        max_num=user.endindex-user.startindex
        val = (userid,max_num,user.startindex)
        sql="SELECT title,authors,issuedate,returndate FROM usage join bookmaster on usage.accno=bookmaster.accno where userid=%s order by issuedate limit %s offset %s"
        mycursor.execute(sql, val,)
        myresult = mycursor.fetchall()
        mycursor.close()
        mydb.close()
        response.set_cookie(key="Manish",value=login.setcookie(login.verifyuser(Manish)), httponly=True,secure=settings.SECURITYHHTPS, samesite=settings.SAMESITE)
        if myresult:
            return(myresult)
        else:
            return
    except:
        return({"msg":"Connection error"})

class Bookusagereport(BaseModel):
    startindex:int
    endindex:int

@router.post("/Book_Usage_report")
def personalreport(response:Response,user:Bookusagereport,Manish:Optional[str]=Cookie(None)):
    role=login.getrole(Manish)
    
    if role!=1:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"invalid user")
    if(user.endindex-user.startindex>10):
        return({"msg":"Maximum 10 books can be fetched"})
    if(user.startindex<0):
        return({"msg":"Start index minimum value should be zero"})
    
    if(user.endindex<user.startindex):
        return({"msg":"end index should be greater than start index"})

    try:

        mydb = psycopg2.connect(
        host=settings.HOST_NAME,
        user=settings.USER_NAME,
        password=settings.USER_PASSWORD,
        database=settings.DATABASE_NAME,
        cursor_factory=RealDictCursor
        )
        mycursor = mydb.cursor()
        max_num=user.endindex-user.startindex
        val = (max_num,user.startindex)
        sql="select title,authors,count(title) from usage join bookmaster on bookmaster.accno=usage.accno  group by title, authors order by title limit %s offset %s"
        mycursor.execute(sql, val,)
        myresult = mycursor.fetchall()
        mycursor.close()
        mydb.close()
        response.set_cookie(key="Manish",value=login.setcookie(login.verifyuser(Manish)), httponly=True,secure=settings.SECURITYHHTPS, samesite=settings.SAMESITE)
        if myresult:
            return(myresult)
        else:
            return
    except:
        return({"msg":"Connection error"})

@router.post("/Book_Usage_report_pdf")
def personalreport(response:Response,user:Bookusagereport,Manish:Optional[str]=Cookie(None)):
    role=login.getrole(Manish)
    
    if role!=1:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"invalid user")
    if(user.endindex-user.startindex>10):
        return({"msg":"Maximum 10 books can be fetched"})
    if(user.startindex<0):
        return({"msg":"Start index minimum value should be zero"})
    
    if(user.endindex<user.startindex):
        return({"msg":"end index should be greater than start index"})

    try:

        mydb = psycopg2.connect(
        host=settings.HOST_NAME,
        user=settings.USER_NAME,
        password=settings.USER_PASSWORD,
        database=settings.DATABASE_NAME,
        cursor_factory=RealDictCursor
        )
        mycursor = mydb.cursor()
        max_num=user.endindex-user.startindex
        val = (max_num,user.startindex)
        sql="select title,authors,count(title) from usage join bookmaster on bookmaster.accno=usage.accno  group by title, authors order by title limit %s offset %s"
        mycursor.execute(sql, val,)
        myresult = mycursor.fetchall()
        mycursor.close()
        mydb.close()

        class PDF(FPDF):
            

            # Page footer
            def footer(self):
            # Position at 1.5 cm from bottom
                self.set_y(-15)
                # Arial italic 8
                self.set_font('Arial', 'I', 8)
                # Page number
                self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

            # Instantiation of inherited class
        pdf = PDF(orientation="landscape", format="A4")
        pdf.alias_nb_pages()
            # pdf = pdf(orientation="landscape", format="A4")
            # fpd PDF()
        pdf.add_page()

        pdf.set_font("Times", size=14,style="B")
        pdf.multi_cell(w=0,h=pdf.font_size * 2.5,txt="Book Usage Report",align="C",border = 0,ln=3,max_line_height=pdf.font_size)
        pdf.ln()
        pdf.set_font("Times", size=12,style="B")
        colwidth=[100,80,50,25,22,22,22,22,22]
        pdf.multi_cell(w=colwidth[0],h=pdf.font_size * 4.5,txt=str("Title"),align="C",border = 1,ln=3,max_line_height=pdf.font_size)
        pdf.multi_cell(w=colwidth[1],h=pdf.font_size * 4.5,txt=str("Authors"),align="C",border = 1,ln=3,max_line_height=pdf.font_size)
        pdf.multi_cell(w=colwidth[2],h=pdf.font_size * 4.5,txt=str("Count"),align="C",border = 1,ln=3,max_line_height=pdf.font_size)
        
        pdf.ln()
        for line in myresult:
            pdf.multi_cell(w=colwidth[0],h=pdf.font_size * 4.5,txt=str(line['title']),align="L",border = 1,ln=3,max_line_height=pdf.font_size)
            pdf.multi_cell(w=colwidth[1],h=pdf.font_size * 4.5,txt=str(line['authors']),align="L",border = 1,ln=3,max_line_height=pdf.font_size)
            pdf.multi_cell(w=colwidth[2],h=pdf.font_size * 4.5,txt=f"{line['count']}",align="C",border = 1,ln=3,max_line_height=pdf.font_size)
            pdf.ln()
            
        pdf.ln()
        pdf.output(f"bookusage_{user.endindex}.pdf")
        response.set_cookie(key="Manish",value=login.setcookie(login.verifyuser(Manish)), httponly=True,secure=settings.SECURITYHHTPS, samesite=settings.SAMESITE)
        return FileResponse(f"bookusage_{user.endindex}.pdf",media_type="application/pdf",filename=f"bookusage_{user.endindex}.pdf")
    except:
        return({"msg":"Connection error"})

@router.post("/Book_Usage_report")
def personalreport(response:Response,user:Bookusagereport,Manish:Optional[str]=Cookie(None)):
    role=login.getrole(Manish)
    
    if role!=1:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"invalid user")
    if(user.endindex-user.startindex>10):
        return({"msg":"Maximum 10 books can be fetched"})
    if(user.startindex<0):
        return({"msg":"Start index minimum value should be zero"})
    
    if(user.endindex<user.startindex):
        return({"msg":"end index should be greater than start index"})

    try:

        mydb = psycopg2.connect(
        host=settings.HOST_NAME,
        user=settings.USER_NAME,
        password=settings.USER_PASSWORD,
        database=settings.DATABASE_NAME,
        cursor_factory=RealDictCursor
        )
        mycursor = mydb.cursor()
        max_num=user.endindex-user.startindex
        val = (max_num,user.startindex)
        sql="select title,authors,count(title) from usage join bookmaster on bookmaster.accno=usage.accno  group by title, authors order by title limit %s offset %s"
        mycursor.execute(sql, val,)
        myresult = mycursor.fetchall()
        mycursor.close()
        mydb.close()
        response.set_cookie(key="Manish",value=login.setcookie(login.verifyuser(Manish)), httponly=True,secure=settings.SECURITYHHTPS, samesite=settings.SAMESITE)
        if myresult:
            return(myresult)
        else:
            return
    except:
        return({"msg":"Connection error"})

@router.post("/Book_Usage_report_pdf")
def personalreport(response:Response,user:Bookusagereport,Manish:Optional[str]=Cookie(None)):
    role=login.getrole(Manish)
    
    if role!=1:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"invalid user")
    if(user.endindex-user.startindex>10):
        return({"msg":"Maximum 10 books can be fetched"})
    if(user.startindex<0):
        return({"msg":"Start index minimum value should be zero"})
    
    if(user.endindex<user.startindex):
        return({"msg":"end index should be greater than start index"})

    try:

        mydb = psycopg2.connect(
        host=settings.HOST_NAME,
        user=settings.USER_NAME,
        password=settings.USER_PASSWORD,
        database=settings.DATABASE_NAME,
        cursor_factory=RealDictCursor
        )
        mycursor = mydb.cursor()
        max_num=user.endindex-user.startindex
        val = (max_num,user.startindex)
        sql="select title,authors,count(title) from usage join bookmaster on bookmaster.accno=usage.accno  group by title, authors order by title limit %s offset %s"
        mycursor.execute(sql, val,)
        myresult = mycursor.fetchall()
        mycursor.close()
        mydb.close()

        class PDF(FPDF):
            

            # Page footer
            def footer(self):
            # Position at 1.5 cm from bottom
                self.set_y(-15)
                # Arial italic 8
                self.set_font('Arial', 'I', 8)
                # Page number
                self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')

            # Instantiation of inherited class
        pdf = PDF(orientation="landscape", format="A4")
        pdf.alias_nb_pages()
            # pdf = pdf(orientation="landscape", format="A4")
            # fpd PDF()
        pdf.add_page()

        pdf.set_font("Times", size=14,style="B")
        pdf.multi_cell(w=0,h=pdf.font_size * 2.5,txt="Book Usage Report",align="C",border = 0,ln=3,max_line_height=pdf.font_size)
        pdf.ln()
        pdf.set_font("Times", size=12,style="B")
        colwidth=[100,80,50,25,22,22,22,22,22]
        pdf.multi_cell(w=colwidth[0],h=pdf.font_size * 4.5,txt=str("Title"),align="C",border = 1,ln=3,max_line_height=pdf.font_size)
        pdf.multi_cell(w=colwidth[1],h=pdf.font_size * 4.5,txt=str("Authors"),align="C",border = 1,ln=3,max_line_height=pdf.font_size)
        pdf.multi_cell(w=colwidth[2],h=pdf.font_size * 4.5,txt=str("Count"),align="C",border = 1,ln=3,max_line_height=pdf.font_size)
        
        pdf.ln()
        for line in myresult:
            pdf.multi_cell(w=colwidth[0],h=pdf.font_size * 4.5,txt=str(line['title']),align="L",border = 1,ln=3,max_line_height=pdf.font_size)
            pdf.multi_cell(w=colwidth[1],h=pdf.font_size * 4.5,txt=str(line['authors']),align="L",border = 1,ln=3,max_line_height=pdf.font_size)
            pdf.multi_cell(w=colwidth[2],h=pdf.font_size * 4.5,txt=f"{line['count']}",align="C",border = 1,ln=3,max_line_height=pdf.font_size)
            pdf.ln()
            
        pdf.ln()
        pdf.output(f"bookusage_{user.endindex}.pdf")
        response.set_cookie(key="Manish",value=login.setcookie(login.verifyuser(Manish)), httponly=True,secure=settings.SECURITYHHTPS, samesite=settings.SAMESITE)
        return FileResponse(f"bookusage_{user.endindex}.pdf",media_type="application/pdf",filename=f"bookusage_{user.endindex}.pdf")
    except:
        return({"msg":"Connection error"})

@router.post("/Book_Usage_report_xlsx")
def personalreport(response:Response,user:Bookusagereport,Manish:Optional[str]=Cookie(None)):
    role=login.getrole(Manish)
    
    if role!=1:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail=f"invalid user")
    if(user.endindex-user.startindex>10):
        return({"msg":"Maximum 10 books can be fetched"})
    if(user.startindex<0):
        return({"msg":"Start index minimum value should be zero"})
    
    if(user.endindex<user.startindex):
        return({"msg":"end index should be greater than start index"})

    try:

        mydb = psycopg2.connect(
        host=settings.HOST_NAME,
        user=settings.USER_NAME,
        password=settings.USER_PASSWORD,
        database=settings.DATABASE_NAME,
        cursor_factory=RealDictCursor
        )
        mycursor = mydb.cursor()
        max_num=user.endindex-user.startindex
        val = (max_num,user.startindex)
        sql="select title,authors,count(title) from usage join bookmaster on bookmaster.accno=usage.accno  group by title, authors order by title limit %s offset %s"
        mycursor.execute(sql, val,)
        myresult = mycursor.fetchall()
        mycursor.close()
        mydb.close()

        workbook = Workbook()
        sheet = workbook.active
        sheet["C1"] = f"Book Usage Report"
        sheet["B2"] = f"Title"
        sheet["C2"] = f"Authors"
        sheet["D2"] = "Count"
            
        i=3
        for line in myresult:
            sheet[f"B{i}"] = line['title']
            sheet[f"C{i}"] = line['authors']
            sheet[f"D{i}"] = line['count']
            i=i+1
            
        workbook.save(f"bookusage_{user.endindex}.xlsx")
        response.set_cookie(key="Manish",value=login.setcookie(login.verifyuser(Manish)), httponly=True,secure=settings.SECURITYHHTPS, samesite=settings.SAMESITE)
        return FileResponse(f"bookusage_{user.endindex}.xlsx",media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",filename=f"bookusage_{user.endindex}.xlsx")
    except:
        return({"msg":"Connection error"})